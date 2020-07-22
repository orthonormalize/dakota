# 20200601:
# gamma main

import numpy as np
import pandas as pd
#import re
import sys
import os
import io
import datetime
from openpyxl import load_workbook

def readQC(inputfile,proc,fieldTable,sheetname=None): # placeholder
    def file2df(inputfile,sheetname=None):
        assert (inputfile), 'missing input file name'
        IFS = inputfile.split('.')
        extension=IFS[-1]
        if (extension=='csv'):
            df = pd.read_csv(inputfile)
        elif (IFS[-1]=='xlsx'):
            with open(inputfile, "rb") as f:
                in_mem_file = io.BytesIO(f.read())
            wb = load_workbook(in_mem_file, read_only=True)
            shNames = wb.sheetnames
            if ((sheetname) and (sheetname not in shNames)):
                raise ValueError('Cannot find sheet name %s in excel file %s' % (sheetname,inputfile))
            elif (sheetname is None):
                if (len(shNames)==1):
                    sheetname = shNames[0]
                else:
                    raise ValueError('Excel file %s has %d sheets. Please specify sheet name to read.' % (inputfile,len(shNames)))
            data = wb[sheetname].values
            try:
                columns = next(data)[0:] # get header line separately
                df = pd.DataFrame(data, columns=columns)
            except StopIteration:
                raise StopIteration('No data in input file %s, sheet %s' % (inputfile,sheetname))
        else:
             raise ValueError('Cannot read input file with extension .%s' % extension)
        df.fillna('', inplace=True)
        df.replace(to_replace='',value=np.nan,inplace=True)   # empty cells were {NaN,None,''} ==> Now all empty cells contain NaN
        return(df)
    
    def rowFiltering(df,proc,fieldTable):
        FT = fieldTable.loc[(fieldTable.PROC==proc) & (fieldTable.Field0.apply(lambda x: len(x)>0))]
        assert all(FT.ifEmpty.isin(['error','filter','ok','okay'])), 'fields: Invalid entry for "ifEmpty", proc %s' % proc
        for f0 in FT.Field0[FT.ifEmpty.isin(['filter'])]:   # 1) apply all row filters
            df = df.loc[(~df[f0].isna())]   
        for f0 in FT.Field0[FT.ifEmpty.isin(['error'])]:    # 2) assert that all mandatory fields are occupied
            assert (all(~df[f0].isna())), 'Input file contains at least one missing data value in required field %s' % f0
        return(df)
    
    df = file2df(inputfile,sheetname)
    df = rowFiltering(df,proc,fieldTable)
    return(df)
        
def commandLine2Dict(CL):
    # input: list of command line arguments
        # m: mapfile
        # f: rawfile
    # output: dict
    D={}
    CL=CL[::-1]
    key=None
    while(CL):
        s=CL.pop()
        if (s.startswith('-')):
            key = (s[1:] if s[1:] not in D else None)
        else:
            if key:
                D[key]=s
            key=None
    return D

def adjustDtypes_Map(M):
    # update 20200604: this function is probably irrelevant
    # input: Map is entirely string data
    # output: convert datatype of any column deemed necessary
    for (_,row) in M['dtypeMap'].iterrows():
        assert(row['Type'] not in ['date','datetime']) # do these later, if necessary
        if (row['Type']=='bool'):
            M[row['Sheet']][row['Field']] = M[row['Sheet']][row['Field']].astype(int).astype(bool)
        else:
            M[row['Sheet']][row['Field']] = M[row['Sheet']][row['Field']].astype(row['Type'])
    return M

def readAllSheets(xlFile):
    # input: path to MS xlsx file
    # output: dict of dataframes. One df for each sheet within xlsx file
    # all sheets must have a header row
    with open(xlFile, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = load_workbook(in_mem_file, read_only=True)
    D = {}
    for sh in wb.sheetnames:
        #D[sh] = pd.read_excel(mapfile,sheet_name='Params',engine='openpyxl',usecols=[0,1],index_col=0)
        data = wb[sh].values
        try:
            columns = next(data)[0:] # get header line separately
            D[sh] = pd.DataFrame(data, columns=columns)
            D[sh] = D[sh][[c for c in columns if c]]  # rm extra columns (caused by sheets with cell colors / old edits ???)
            D[sh] = D[sh].applymap(lambda x: ('' if (x==None) else str(x)))  # converts ALL data to str. Empty cells are ''
            D[sh] = D[sh].applymap(lambda x: x.strip()) 
        except StopIteration:
            # ignore empty sheets
            pass
    return D

def nestSheets(M,startswith):
    # input: dict of DFs generated from reading all sheets of map file
    # output: smaller dict of DFs, consisting only of the sheets whose names begin with string 'startswith'
        # that string will be chopped off of the new dictionary names
    outD = {}
    for k in M.keys():
        if (k.startswith(startswith)):
            shortName = k[len(startswith):]
            outD[shortName] = M[k]
    return outD

def df2dict(df,keys,fullsheetname):
    # Function: create a dictionary of function mappings from a DF and a list of which fields are keys
    # input: pandas DataFrame
    # output: Triply nested dict
        # lev 1 keys = keys arg
        # lev 2 keys = each other field in the DF
        # lev 3 keys = DF values of the key field. Enforces uniqueness
        # values     = DF values of the field (lev2 key)
    D=dict.fromkeys(keys,0)
    for k1 in keys:
        pS = getattr(df,k1)
        assert pS.is_unique,('Found at least one duplicate value in %s:%s' % (fullsheetname,k1))
        df_k1 = df.set_index(k1,drop=True,inplace=False)
        D[k1] = {col:(df_k1[col].to_dict()) for col in df_k1}
    return D
        
        
def makeDatadicts(M,startswith='data'):
    # input: dict of DFs generated from reading all sheets of map file
        # should contain one sheet called 'dicts' that lists all key fields in format: sheet name, field name
        # optional input argument 'startswith' appends a string to the beginning of each sheet name string
    # generate a hash table from each entry in 'dicts'
        # assert uniqueness of corresponding table/field
    # then output: a dict parent pointing to all keyfields, then valuefields as nested hash tables 
    D = {table:{} for table in getattr(M.get('dicts'),'SHEET').unique()}
    for (_,(sh,key)) in M.get('dicts').iterrows():
        D[sh][key]={}
    for sh in D:
        fullsheetname=startswith+sh
        df = M.get(fullsheetname)
        D[sh] = df2dict(df,D[sh].keys(),fullsheetname)
    return D

def makeParams(M,sheetname='params'):
    return (M.get(sheetname).set_index('Parameter',drop=True)['Value'].to_dict())