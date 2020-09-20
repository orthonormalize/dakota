import numpy as np
import pandas as pd
import re
import os
import io
import datetime
from collections import Counter
from openpyxl import load_workbook
import json


globs = {k:globals()[k] for k in globals() if not (k.startswith('_'))}
import builtins
builts = {s:getattr(builtins,s) for s in builtins.__dir__()}
import dakotaLib as DL
objs_DL = {k:getattr(DL,k) for k in DL.__dir__() if not (k.startswith('_'))}



# Define Constants:
FT_splitch = '_'
defaultDateFormat = '%Y/%m/%d'
empties = {'str':'','int':0,'bool':False,'date':np.datetime64('1970-01-01'),'float':np.nan}
typeFunctions = {'str':str,'int':int,'bool':bool,'float':float,'date':np.datetime64}
setDigits = set([chr(j) for j in range(48,58)])
setAlpha = set([chr(j) for j in range(65,91)] + [chr(j) for j in range(97,123)])
setAlphaUS = setAlpha.copy()
setAlphaUS.update('_')
setAlphaUSDigit = setAlphaUS.copy()
setAlphaUSDigit.update(setDigits)
loopKeywords = ['for'] # not yet configured to handle any other loop constructs
identifierMapper = {'if':'iff'}

loweralphanum = lambda x: ''.join(ch for ch in x.lower() if ch.isalnum())

# Type Converters:
def is_date(x):
    return (isinstance(x,np.datetime64) or (isinstance(x,datetime.date)) or isinstance(
                                            x,pd._libs.tslibs.timestamps.Timestamp))

def forceType(x,targetString,dateformatStr=defaultDateFormat):
    # x === scalar object. type of x must be in {int, float, bool, str, or date}
    # targetString must be in dict typeFunctions
    x = (np.nan if (x=='') else x)
    if (pd.isna(x)):
        return(empties[targetString])
    if (targetString=='date'):
        if (isinstance(x,str)):
            return(np.datetime64(datetime.datetime.strptime(x,dateformatStr)))
        assert is_date(x), 'Cannot convert %s %s to date' % (str(type(x)),str(x))
        return(np.datetime64(x))
    elif (is_date(x) and (targetString=='str')):
        x=pd.to_datetime(x).strftime(dateformatStr)
    elif (isinstance(x,str)):
        x=x.strip()
        x = (False if (x.lower()=='false') else (True if (x.lower()=='true') else x))
    if (targetString!='str'):
        x=float(x)
    if (targetString in ['int','bool']):
        x=round(x)
    return(typeFunctions[targetString](x))

def convertTypePS(ps0,targetString,dateformatStr=defaultDateFormat):
    return ps0.apply(lambda x: forceType(x,targetString,dateformatStr)).astype(typeFunctions[targetString])


# Define lots of object types, with inheritances
    # Not used for data storage, just to define types and relationships

class o4StrJoin:                          # item handle needs conversion from charList to str, via (''.join())
                                                # two options: either (1) LITERAL or (2) IDENTIFIER
    def D_s2o():
        D = dict([(a,oPKMW) for a in setAlphaUS])          # assume alphaUS_identifier is an Attribute until proven otherwise
        D.update(dict([(a,oInt) for a in setDigits]))
        D.update({'-':oInt,'@':oXat,'#':oHashat})
        D.update({k:oGrouper.BG_2_objtype(k) for k in oGrouper.BG}) # FLAG: includes some irrelevant BGs here
        return(D)
    
    def starter2objtype(ch0):
        return(o4StrJoin.D_s2o()[ch0])
    
    def initiation_prechar(self):
        return(self.initiation_prechar)

class oGrouper:                           # any object that is closed by a specific specialCharacter
    groupers_OC = (('()','[]','{}','""',"''"))
    BG = [goc[0] for goc in groupers_OC]
    EG = [goc[1] for goc in groupers_OC]
    def BG_2_objtype(bg):
        D_bg2obj = {'(':oCallable,'[':oList,'{':oDict,"'":oStr1,'"':oStr2}
        return(D_bg2obj[bg])
    def endGrouper():
        return(self.closingChar)

class oAttr:                              # item follows '.'
    pass
class oNonAttr:                           # item occurring at start of expression, or following ',', or following any BG
    pass

class oLiteral(o4StrJoin):                # item that explicitly states its value w/o references (e.g str, int, bool)
    def converter(self):
        return(self.target_type)

class oStr(oLiteral,oGrouper):
    target_type = str   
    initiation_prechar = ''
    keepEntranceCharacter = False
    validContinuationCharacters = None    # FLAG: actually any char should be allowed. But oStr append doesn't check VCC
class oStr1(oStr):
    openingChar="'"
    closingChar="'"
class oStr2(oStr):
    openingChar='"'
    closingChar='"'
class oInt(oLiteral):
    target_type=int
    initiation_prechar = ''              
    keepEntranceCharacter = True 
    validContinuationCharacters = setDigits
class oIntAttr(oInt,oAttr):               # index into list, or dict key (if key is type int) 
    pass
class oBool(oLiteral):
    target_type=bool
class oNone(oLiteral):
    target_type=lambda x,y=None: None     # why on earth does this one require two args to work ??????

class oIdent(o4StrJoin):                  # {@,#,P,K,M,W} {xat,hashat,property,key(dict),method,callable keyword}
    invalidLowers = {'false':(oBool(),False),'true':(oBool(),True),'none':(oNone(),None)}
class oXat(oIdent):                       # {@}
    initiation_prechar = '0'
    keepEntranceCharacter = False
    validContinuationCharacters = setDigits
class oHashat(oIdent):                    # {#}
    initiation_prechar = ''
    keepEntranceCharacter = False
    validContinuationCharacters = (setAlphaUSDigit | {':'})
class oPKMW(oIdent):                      # {P,K,M,W}: any identifier that starts with a letter or an underscore
    initiation_prechar = ''
    keepEntranceCharacter = True
    validContinuationCharacters = setAlphaUSDigit
class oKeyword(oPKMW,oNonAttr):           # string representing the keyword for a callable kwarg 
    pass
class oPKM(oPKMW,oAttr):                  # {P,K,M} {obj property, dict key, method (either instanceM, or staticM from a lib)}
    pass

class oInternal:                          # object type is internal only (i.e. never visible at mT[-1])
    def obj2postobj(inclass):
        D_o2po = {oCallable:oPostCallable,oList:oPostList,oDict:oPostDict}
        # FLAG: note: this dictionary mapping could fail if any of D_o2po keys ever become superclasses
        return(D_o2po[inclass])
class oList(oInternal,oGrouper):
    openingChar='['
    closingChar=']'
class oDict(oInternal,oGrouper): # oDict is not operational yet
    openingChar='{'
    closingChar='}'
class oCallable(oInternal,oGrouper,oPKM,oAttr):  # either instance method (df.sort_values()) or static method (pd.concat())
    openingChar='('
    closingChar=')'
    
class oPre:
    pass
class oPreAttr(oPre):                     # invoked after '.'
    pass
class oPreNonAttr(oPre):                  # invoked (1) at $ (2) after ',' or (3) after any BG
    pass

class oPost:
    pass
class oPostList(oPost):                   # invoked after ']' 
    pass
class oPostCallable(oPost):               # invoked after ')'
    pass
class oPostDict(oPost):                   # invoked after '}' (placeholder)
    pass
class oPostStr(oPost,oLiteral):           # invoked after endstring character
    target_type=str


def getter(parent,item):
    item = identifierMapper.get(item,item)
        # ALSO CHECK blacklist funcs
    
    if ((isinstance(parent,dict)) and (item in parent)):
        return dict.get(parent,item)
    elif (isinstance(item,int)):
        return parent[item]
    elif (hasattr(parent,item)):
        return getattr(parent,item)
    else:  # hack!    for: executeHashat() where fT0 = 'str.contains' from pandas
        assert ((isinstance(item,str)) and ('.' in item)), 'cannot find item %s(%s) in parent %s' % \
                                                                (item.__class__,item,parent)
        chain=item.split('.')
        return(getter(getter(parent,chain[0]),'.'.join(chain[1:])))
    

class Instruction:
    def __init__(self,*args,**kwargs):
        self.a=args
        self.k=kwargs
        self.X=self.k['X']
        self.procname=self.k['procname']

    def prop2typeparse(self,nameof_targetDF,propName,splitchar=FT_splitch):
        F=self.X['fields']
        rowMatched = F.loc[(F['proc']==self.procname) & (F['object']==nameof_targetDF) & (F['property']==propName)]
        assert (len(rowMatched)==1), '"Fields" has %d matches for (proc,obj,prop) = (%s,%s,%s)' % (
                                                                    len(rowMatched),self.procname,nameof_targetDF,propName)
        row = next(rowMatched.itertuples())
        sType = row.type
        L = re.split(re.escape(splitchar),sType)
        return(L)
    
    def prop2typestr(self,nameof_targetDF,propName,splitchar=FT_splitch):
        L = self.prop2typeparse(nameof_targetDF,propName,splitchar=FT_splitch)
        return(L[0])
    
    def prop2typeconverter(self,nameof_targetDF,propName,mode_ico='c',splitchar=FT_splitch):
        # propName: official name of pd Series object, from X['fields'].property
        # mode_ico: one of {'i','c','o'}, for {'input','computations','output'}.
            # Allows objects to be represented in different formats for I/O than they are in internal operations 'C'
            # mode 'O' will force conversion of date objects to strings before outputting
        # Date Formats are the only special formats fully customizable yet by this function (2020 August)
        # OUTPUT: lambda function, taking pd Series arg, that applies the correct type conversion to the arg
        L = self.prop2typeparse(nameof_targetDF,propName,splitchar=FT_splitch)
        targetString = L[0]
        dateformatString = None
        for s in L[1:]:
            assert (s[0] in 'ico')
            assert (s[1] == '=')
            if s.startswith(mode_ico):
                dateformatString = s[2:]
        if (targetString=='date'):
            if (dateformatString is None):
                dateformatString = defaultDateFormat
            if (mode_ico=='o'):
                targetString='str' # if output mode, override 'date' and force 'str'. it will be converted with strftime
        return (lambda pandasSeries: convertTypePS(pandasSeries,targetString,dateformatStr=dateformatString))
    
    def jsonDump(self,priority,obj0,filebase,includeTimeString=True):
        if (priority<=self.X['params']['picklosity']):
            if (isinstance(obj0,dict) and ('obj' in obj0) and ('task' in obj0) and \
                                    (isinstance(obj0['obj'],pd.Series) or isinstance(obj0['obj'],pd.DataFrame))):
                try:
                    DLTS = DL.timeString(includeTimeString)
                    with open(filebase + DLTS + "_0task.json", "w" ) as outfile:
                        json.dump(obj0['task'],outfile)
                    with open(filebase + DLTS + "_1obj.json", "w" ) as outfile:
                        obj0['obj'].to_json(outfile)
                except:
                    pass
            else:
                if (isinstance(obj0,pd.Series) or isinstance(obj0,pd.DataFrame)):
                    try:
                        with open(filebase+DL.timeString(includeTimeString)+".json","w") as outfile: 
                            obj0.to_json(outfile)  # pd Series or DataFrame
                    except:
                        pass
                else:    
                    try:
                        with open(filebase + DL.timeString(includeTimeString) + ".json", "w" ) as outfile:
                            json.dump(obj0,outfile)
                    except:
                        pass
            
    def getObj0(self,obj2find):
        # obj2find === string, name of desired object
        # return a dict pointing to the obj2find's parent 
            # Search priority: {X, globs, builts, objs_DL, self.__dir__()}
            # This function can only return 1 of 5 dicts: {X, globs, builts, objs_DL, instanceAttrs}
                # Not an instance method. All instance data must reach getObj in one of two ways:
                    # through X['@'], OR
                    # through Hashats (e.g. X['data'])
        obj2find = identifierMapper.get(obj2find,obj2find)
        for searchSpace in [self.X,globs,builts,objs_DL]:
            if (obj2find in searchSpace):
                return(searchSpace)
        # if not yet found, it had better be in self.__dir__() (e.g. it's a direct call of an Instruction instance method)
        instanceAttrs = {k:getattr(self,k) for k in self.__dir__() if not (k.startswith('_'))}
        assert (obj2find in instanceAttrs), 'Cannot find object: %s' % obj2find
        return(instanceAttrs)
    
    def getObj(self,s,obj0=None,ifEmpty=None,ignoreLevels=0):
        # s === string that refers to a chain that can be resolved into objects and/or methods
        # Find and return the corresponding object within the dataStructure.
        # s can be empty. If so, return ifEmpty
        
        def retrieveArgumentName_Callable():
            D = callableDictStack[-1]
            assert (D['cur'] is not None), 'only keyword args may follow a keyword arg'
            D['keyList'].append(D['cur'])
            if (D['kwQ']): # kwarg:
                D['cur']=None
            else: # numbered arg:
                D['cur']+=1
                
        def executeCallable():
            # input state: (uses vars from scope:getObj7)
                # callableDictStack[-1] must be dict {kwQ,keyList,cur}. Keywords obtained here. ===> D
                # final two elements of container[-1] should be:
                    # method (either boundMethod of instance, or staticMethod of another class): ===> objFUNK
                    # list of arguments to be passed into the callable: ===> argList
            # output state:
                # callableDictStack has popped off one used item
                # container[-1]: two items popped off (objFUNK, argList); then one item pushed (output from method call)
            # This function DOES NOT change len(container), DOES NOT modify mT at all
            D = callableDictStack.pop()
            argList = container[-1].pop()
            objFUNK = container[-1].pop()
            assert (len(argList)==len(D['keyList'])), 'length mismatch: keys:%d, args:%d' % (len(D['keyList']),len(argList))
            dex=0
            while dex in D['keyList']:
                dex+=1
            A = argList[:dex]
            K = dict(zip(D['keyList'][dex:],argList[dex:]))
            container[-1].append(objFUNK(*A,**K))
        
        def resolve():
            # i.e. stepOut
            # input state: ch is either in {EG, ',', '.'}, or ch is '(' interrupting an identifier 
            # output: resolved container[-1] into previous level of stack 
            # function: 
                # 1) Pop container[-1]. Assemble: Combine / interpret / convert / etc., if necessary. 
                # 2) Drop as one object into the parent level of the container stack.
                # 3) Attribute processing: (a) get attr if PKM, or (b) execute if postcallable
                        # (c) if neither, but destination is callable: update callableDictStack
                # 4) Pop mT
            R = container.pop()
            if (isinstance(mT[-1],o4StrJoin)):
                R=''.join(R)
                if (isinstance(mT[-1],oIdent) and (R.lower() in mT[-1].invalidLowers)): 
                    (mT[-1],R) = mT[-1].invalidLowers[R.lower()]  # hack: detects/typeConverts {booleans, None}
                if (isinstance(mT[-1],oLiteral)):
                    R = mT[-1].converter()(R)
                elif (isinstance(mT[-1],oXat)):
                    R = (self.X['@'][int(R)])
                elif (isinstance(mT[-1],oHashat)):
                    R = self.executeHashat(R)
            container[-1].append(R)
            if (isinstance(mT[-1],oAttr) or isinstance(mT[-1],oPKMW)):
                attribute = container[-1].pop()
                attribute = identifierMapper.get(attribute,attribute)
                if (isinstance(mT[-1],oAttr)): # parent is already present in container
                    baseObj = (container[-1].pop())
                else: # This is an orphan identifier, so need to search for parent:
                    baseObj = self.getObj0(attribute)
                    if (isinstance(mT[-2],oCallable)): 
                        # ancestor of orphan had never been written as callableArg, so CDS still needs update!
                        retrieveArgumentName_Callable()
                container[-1].append(getter(baseObj,attribute))
            elif (isinstance(mT[-1],oPostCallable)):
                executeCallable()
            elif (isinstance(mT[-2],oCallable)):
                retrieveArgumentName_Callable()
            mT.pop()
            
        def stepIn(initiateType):
            mT.append(initiateType())
            container.append([])
               
        assert ((isinstance(s,str))), 's must be type <str>'
        if (not(s)):
            return(ifEmpty)
        if (ignoreLevels):
            ccc = Counter(s)
            numLevels = ccc.get('.',0)
            if (numLevels==(ignoreLevels-1)):
                if (obj0 is None):
                    return(self.getObj0(s.split('.')[0]))  # possible cryptic error/non-error if s is malformed
                else:
                    return(obj0)
            assert (numLevels>=ignoreLevels), \
                                'Error: ignoreLevels=%d but input string has %d levels' % (ignoreLevels,numLevels)
            assert (not(any([ch in (set('()[]{},="'+"'")) for ch in ccc]))), \
                                'specChars only permitted when ignoreLevels=0'
            s = '.'.join(s.split('.')[:-ignoreLevels])
            assert (s), 'Bad input string'  # possible cryptic error/non-error if s is malformed (e.g. if s=='..')
            
        # 0) initialize:
        (mT,container,callableDictStack) = ([None],[[obj0]],[])
        if (obj0):
            stepIn(oPreAttr)
        else:
            stepIn(oPreNonAttr)
        
        # 1) process each char:
        for ch in s:
            # update all the asserts here to catch more illegal input
            # A)
            if (isinstance(mT[-1],oStr)):
                if (ch==mT[-1].closingChar):
                    mT[-1] = oPostStr()
                else:
                    container[-1].append(ch)
            # B)
            elif (ch=='.'):
                # assert valid?
                resolve()
                stepIn(oPreAttr)
            # C)
            elif (ch==','):
                # assert valid?
                resolve()
                stepIn(oPreNonAttr)
            # D)
            elif (ch=='='):
                assert ((isinstance(callableDictStack[-1]['cur'],int)) or (callableDictStack[-1]['cur'] is None)), \
                                "Illegal string %s: one argument cannot be assigned to more than one keyword" % s
                assert (isinstance(mT[-1],oPKMW)), 'Attempted to create keyword name %s, but it is not an Identifier' % keyword
                assert (isinstance(mT[-2],oCallable)), 'Attempted keyword assignment %s but target is not a Callable' % keyword
                keyword=(''.join(container.pop()))
                container.append([])  # not a stepOut: restart a fresh container to hold the value of this argument
                callableDictStack[-1].update({'kwQ':True,'cur':keyword}) # stage name of keyword into 'cur'
                mT[-1] = oPreNonAttr()
            # E)
            elif ((ch in oGrouper.EG) and (ch not in oGrouper.BG)):
                assert (isinstance(mT[-2],oGrouper)), 'End groupers are not valid when mT[-2] is %s' % str(mT[-2]) 
                assert (ch==mT[-2].closingChar), 'Character "%s" is not a valid closer for %s' % (ch,str(mT[-2]))
                if (isinstance(mT[-1],oPre)): # don't resolve() if contents are still empty! Just stepOut via pops alone
                    mT.pop()
                    container.pop()
                else:
                    resolve()
                assert (isinstance(mT[-1],oInternal)), 'Object closed by end grouper must be oInternal, not %s' % str(mT[-1])
                mT[-1] = oInternal.obj2postobj(type(mT[-1]))()
            # F)
            elif (isinstance(mT[-1],oPre)):
                if ((ch in oGrouper.BG) and (ch not in oGrouper.EG)): # Q: better way to reject stringGroupers here?
                    assert (ch!='('), "Literal Tuples Disabled. Callable must interrupt oIdent, not %s" % str(mT[-1])
                    mT[-1] = oGrouper.BG_2_objtype(ch)()
                    stepIn(oPreNonAttr)
                else:
                    new_mT_candidate = o4StrJoin.starter2objtype(ch)()
                    if (isinstance(mT[-1],oPreAttr)):
                        # (if object is dot-accessed, it must be an instance of oAttr. Modify mT_candidate accordingly!)
                        # Is there a better way to solve this problem thru class relationships rather than bruteForce here?
                        if (isinstance(new_mT_candidate,oInt)):   # e.g. 'L.0'
                            new_mT_candidate = oIntAttr()      
                        elif (isinstance(new_mT_candidate,oPKMW)):   # e.g. 'pd.concat'
                            new_mT_candidate = oPKM()
                    mT[-1] = new_mT_candidate
                    container[-1].append((ch) if (mT[-1].keepEntranceCharacter) else mT[-1].initiation_prechar)
            # G)
            elif (ch in oGrouper.BG): # i.e. '(' entering a callable
                assert (ch=='('), 'Illegal character "%s" interrupting %s'  %  (ch,str(mT[-1]))
                resolve()
                stepIn(oCallable)
                stepIn(oPreNonAttr)
                callableDictStack.append({'kwQ':False,'keyList':[],'cur':0})
            # H)
            else:
                assert (ch in mT[-1].validContinuationCharacters), 'Illegal char "%s" in %s' % (ch,str(mT[-1]))
                container[-1].append(ch)
        
        # Final resolve():
        resolve()
        return(container[-1][-1])
        
    
class Statement(Instruction):
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        self.TASK = self.k['TFT'].TASK
        self.GET = self.k['TFT'].GET
        self.SET = self.k['TFT'].SET
        
    def readQC(self,inputfile,nameof_targetDF=None,sheetname=None,**kwargs):
        # inputfile: str: name of CSV or XLSX file to read
        # nameof_targetDF:  str: name (within X) of DataFrame object where result will be stored
            # this will determine what row subset from X['fields'] gets used
            # if omitted, use the variable name string from self.SET
        # sheetname: str: if inputfile is XLSX, may need to specify sheetname if file contains more than one sheet
        
        def file2df(inputfile,sheetname=None):
            assert (inputfile), 'missing input file name'
            IFS = inputfile.split('.')
            extension=IFS[-1]
            if (extension=='csv'):
                dtype=(kwargs.pop('dtype') if ('dtype' in kwargs) else str)
                df = pd.read_csv(inputfile, dtype=dtype, **kwargs)
            elif (extension=='xls'):
                dtype=(kwargs.pop('dtype') if ('dtype' in kwargs) else str)
                df = pd.read_excel(inputfile, dtype=dtype, **kwargs)
            elif (extension=='xlsx'):
                with open(inputfile, "rb") as f:
                    in_mem_file = io.BytesIO(f.read())
                wb = load_workbook(in_mem_file, read_only=True, **kwargs)
                shNames = wb.sheetnames
                if ((sheetname) and (sheetname not in shNames)):
                    raise ValueError('Cannot find sheet name %s in excel file %s' % (sheetname,inputfile))
                elif (sheetname is None):
                    if (len(shNames)==1):
                        sheetname = shNames[0]
                    else:
                        raise ValueError('Excel file %s has %d sheets. Specify sheet to read.' % (inputfile,len(shNames)))
                data = wb[sheetname].values
                try:
                    columns = next(data)[0:] # get header line separately
                    df = pd.DataFrame(data, columns=columns)
                except StopIteration:
                    raise StopIteration('No data in input file %s, sheet %s' % (inputfile,sheetname))
            else:
                 raise ValueError('Cannot readQC input file with extension .%s' % extension)
            df.fillna('', inplace=True)
            df.replace(to_replace='',value=np.nan,inplace=True)   # all empty cells are np.nan after this step
            return(df)

        def rowFiltering(df,proc,nameof_targetDF,fieldTable):
            FT = fieldTable.loc[(fieldTable.proc==proc) & (fieldTable.object==nameof_targetDF) &
                                  (fieldTable.field0.apply(lambda x: len(x)>0))]
            assert all(FT.ifEmpty.isin(['error','filter','ok','okay'])),'fields: Invalid entry: "ifEmpty": proc %s, obj %s' % (
                                                                                                proc,nameof_targetDF)
            for f0 in FT.field0[FT.ifEmpty.isin(['filter'])]:   # 1) apply all row filters
                df = df.loc[(~df[f0].isna())]   
            for f0 in FT.field0[FT.ifEmpty.isin(['error'])]:    # 2) assert that all mandatory fields are occupied
                assert (all(~df[f0].isna())), 'Input file contains at least one missing data value in required field %s' % f0
            return(df)

        def constructDF(df0,proc,nameof_targetDF,fieldTable):
            FT_direct = fieldTable.loc[(fieldTable.proc==proc) & (fieldTable.object==nameof_targetDF) & (
                                       fieldTable.field0.apply(lambda x: len(x)>0))]
            FT_full = fieldTable.loc[(fieldTable.proc==proc) & (fieldTable.object==nameof_targetDF)]
            myField_2_rawField = {FT_direct.property[row]:FT_direct.field0[row] for row in FT_direct.index}
            df1 = pd.DataFrame({f:df0[myField_2_rawField[f]] for f in FT_direct.property},columns=list(FT_full.property))
            return(df1)

        def convert_dtypes(df1,proc,nameof_targetDF,fieldTable):
            FT_direct = fieldTable.loc[(fieldTable.proc==proc) & (fieldTable.object==nameof_targetDF) & (
                                       fieldTable.field0.apply(lambda x: len(x)>0))]
            FT_full = fieldTable.loc[(fieldTable.proc==proc) & (fieldTable.object==nameof_targetDF)]
            for row in FT_full.index:
                column = FT_full.loc[row,'property']
                field0 = FT_full.loc[row,'field0']
                if (field0):
                    df1[column] = self.prop2typeconverter(nameof_targetDF,column,mode_ico='i')(df1[column])
                else:
                    df1[column] = self.prop2typeconverter(nameof_targetDF,column,mode_ico='i')(
                        pd.Series([empties[self.prop2typestr(nameof_targetDF,column)]]*len(df1),
                                  index=df1.index))
            return(df1)

        if (nameof_targetDF is None):
            nameof_targetDF = self.SET
        proc=self.procname
        fieldTable = self.X['fields'][self.X['fields']['proc']==self.procname]
        df0 = file2df(inputfile,sheetname)
        df0 = rowFiltering(df0,proc,nameof_targetDF,fieldTable)
        df1 = constructDF(df0,proc,nameof_targetDF,fieldTable)
        df1 = convert_dtypes(df1,proc,nameof_targetDF,fieldTable)
        return(df1)

    def write_csv(self,df_in,outfilename,**kwargs):
        # wrapper for pd.to_csv
        # purpose: 1) downselect columns for output, according to X.fields
                #  2) standardize all data formats (e.g. date display format)
        # if nameofDF not given, pull the correct string directly from first varname in self.GET
    
        # 1) Find matching proc and object:
        nameofDF = ((kwargs.pop('nameofDF')) if ('nameofDF' in kwargs) else (None))
        assert (df_in is self.X['@'][0]), 'First argument to write_csv() must be GET[0]'
        fieldsP = self.X['fields'][self.X['fields']['proc']==self.procname]
        uObj = fieldsP.object.unique()
        if (nameofDF is None):
            candidate = self.GET.split(',')[0]
            nameofDF = (candidate if (candidate in uObj) else None)
        if (nameofDF is None):
            # Allow a match if there is a fields template with identical columns, albeit under a different name:
            for candidate in uObj:
                if (set(fieldsP[fieldsP.object==candidate].property.unique()) == set(df_in.columns)):
                    nameofDF = candidate
                    break
        assert (nameofDF is not None), 'Cannot find a Fields template for dataframe %s' % (self.GET.split(',')[0])
        assert (set(fieldsP[fieldsP.object==nameofDF].property.unique()) == set(df_in.columns)), (
                            'write_csv: failed to match columns for %s and %s' % (nameofDF, self.GET.split(',')[0]))
        fieldsPO = fieldsP[fieldsP.object==nameofDF]
        # 2) Discards and reFormats to match template output style:
        df_out = df_in.copy()
        df_out = df_out[list(fieldsPO.property.loc[~fieldsPO.discard.isin(set(['1',1,True]))])] # discards and sorts
        for col in df_out.columns:
            df_out.loc[:,col] = self.prop2typeconverter(nameofDF,col,mode_ico='o')(df_out[col])
        df_out.to_csv(outfilename,**kwargs)

    def executeHashat(self,tablename_seq):
        # always operate on X['@'][0], which is assumed to be a pd DataFrame
        # tablename should also be found as a key of X
            # Table needs fields: [PROC,SEQ,FILTER_PROPERTY,FILTER_TYPE,FILTER_VALUE,TARGET_PROPERTY,TARGET_VALUE]
        # if there is a ':' within tablename_seq:   use seq to further select subset of rows from computes
        assert (len(self.X['@'])==1) and (isinstance(self.X['@'][0],pd.DataFrame)), \
                        'Hashat function %s takes exactly one argument which must be a pandas DataFrame' % tablename_seq
        df = self.X['@'][0]
        TS = tablename_seq.split(':')
        assert (len(TS) <= 2), 'Illegal Hashat construction: too many colons'
        (tablename,seq) = ((TS[0],int(TS[1])) if (len(TS)==2) else (TS[0],None))
        assert (tablename in self.X), 'Hashat function name not found'
        table = self.X[tablename]
        # adjust table:   str replacements in col FILTER_TYPE:
        table.FILTER_TYPE = table.FILTER_TYPE.replace({'contains':'str.contains'})
        (pat,repl) = (r"^(([lg][et])|(eq)|(ne))$", lambda m: '__'+m.group(0)+'__')  # comparison operators
        table.FILTER_TYPE = table.FILTER_TYPE.str.replace(pat,repl)
        print(list(table.FILTER_TYPE))
        
        # Process df:
        # 1) downselect rows in df via self.procname and seq
        tableKeeps = (table.PROC==self.procname)
        if (seq is not None):
            tableKeeps = (tableKeeps & (table.SEQ.astype(int) == seq))
        print('keeping table rows:')
        print(tableKeeps)
        hashat = table[tableKeeps]
        
        # 2) Form joined table:
        fieldsInfo = self.X['fields'][self.X['fields']['proc']==self.procname]
        hashat_J = (
            hashat.reset_index()
                .merge(fieldsInfo.add_suffix('_fil'), how="left",
                           left_on='FILTER_PROPERTY',right_on='property_fil',validate="m:1")
                .merge(fieldsInfo.add_suffix('_tar'), how="left",
                           left_on='TARGET_PROPERTY',right_on='property_tar',validate="m:1")
                .set_index('index')
        )  
        print('hashat_J:')
        print(hashat_J)
        
        # 3) Process Hashat, row by row:
        print(len(df))
        for row in hashat_J.itertuples():
            self.X['@'] = [df]
            print()
            print('%02d' % row.Index)
            print(row)
            # 2A) Compute df rowmask:
            if (not(row.FILTER_PROPERTY)): # no filter ==> calculate all rows
                print('FP WAS EMPTY')
                rowmask = pd.Series(True,index=df.index)
            else:
                (fP0,fT0,fV0) = (row.FILTER_PROPERTY,row.FILTER_TYPE,row.FILTER_VALUE)
                filter_is_string = (self.prop2typestr(row.object_fil,fP0)=='str')
                fV1 = self.prop2typeconverter(row.object_fil,fP0,mode_ico='c')(pd.Series([fV0])).iloc[0]
                if (filter_is_string):
                    rowmask = getter(getter(df,fP0).str.lower(), fT0)(fV1.lower())  # case-insensitivity is hardWired
                else:
                    rowmask = getter(getter(df,fP0),fT0)(fV1)
            # 2B) Calculate df column:
            self.X['@'] = [df.loc[rowmask]]
            print((np.count_nonzero(rowmask),len(rowmask)))
            if (np.count_nonzero(rowmask)==1642):
                print([kk for kk in rowmask.index if not(rowmask[kk])])
            if(any(fieldsInfo.property==row.TARGET_PROPERTY)):
                print((type(row.TARGET_CALCULATION),row.TARGET_CALCULATION))
                temp_ps0 = pd.Series(self.getObj(row.TARGET_CALCULATION),index=df[rowmask].index)
                temp_ps1 = self.prop2typeconverter(row.object_tar,row.TARGET_PROPERTY,mode_ico='c')(temp_ps0)
                df.loc[rowmask,row.TARGET_PROPERTY] = temp_ps1
            else: # non-calculation. must be either "remove" or "error"
                if (row.TARGET_PROPERTY=="remove"):
                    print('removing %d rows' % np.count_nonzero(rowmask))
                    df = df.loc[~rowmask]
                else:
                    assert (row.TARGET_PROPERTY=="error"), "Invalid value for hashat:TARGET_PROPERTY: %s"%row.TARGET_PROPERTY
                    assert (not(any(rowmask))), 'Hashat error caught at least one row of df'
            self.jsonDump(3,rowmask,"__dakson_hashat_rowmask_" + ('%02d'%row.Index))
            self.jsonDump(3,df,"__dakson_hashat_df_" + ('%02d'%row.Index))
            print(len(df))
        print()
        return(df)
    
    def execute(self):
        print()
        print()
        print('placeholder: begin executing ' + self.TASK)
        self.X['@'] = [self.getObj(ss,obj0=self.X) for ss in self.GET.split(',')]
        print('finished Statement.GET.    Here is @:')
        print(self.X['@'])
        self.X['#'] = self.getObj(self.TASK)
        print('finished Statement.TASK.    Here is the result:')
        print(self.X['#'])
        
        DiagnosisDictionary = {'task':self.TASK,'obj':self.X['#']}
        self.jsonDump(2,DiagnosisDictionary,"__dakson_s_")
        
        if (self.SET):
            parent = self.getObj(self.SET,obj0=self.X,ignoreLevels=1)
            namesPA = self.SET.split('.')
            (parentName,attrName) = ('.'.join(namesPA[:-1]),namesPA[-1])
            attrName = self.SET.split('.')[-1]
            if (isinstance(parent,pd.DataFrame) and isinstance(self.X['#'],pd.Series)):
                XFP = self.X['fields'][self.X['fields'].proc==self.procname]
                if ((parentName in XFP.object.unique()) and (attrName in XFP[XFP.object==parentName].property.values)):
                    # parent has a data type schema defined that we must honor:
                    self.X['#'] = self.prop2typeconverter(parentName,attrName,mode_ico='c')(self.X['#'])
                self.X['#'].rename(attrName,inplace=True)
                if (self.X['#'].name in parent.columns):
                    parent.update(self.X['#'])
                else:
                    assert self.X['#'].index.equals(parent.index), 'assignment failed: series and DF indices do not match'
                    parent[self.X['#'].name] = self.X['#']
            else:
                _ = ((setattr(parent,attrName,self.X['#'])) if (hasattr(parent,'__getattr__'))
                                                                   else (parent.__setitem__(attrName,self.X['#'])))
        print('placeholder: done executing ' + self.TASK)
    
class Loop(Instruction):
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        self.body = InstructionList(self.X,self.procname,self.k['body'])
        self.controlString=self.k['controlString']
        conList = self.controlString.split(' ')
        assert (len(conList)==4), 'Bad control string: "%s". Control string must contain exactly four words'
        (self.keyword, self.loopvar, mustbe_in, self.iterable) = conList
        self.iterable = ((self.iterable[:-1]) if (self.iterable.endswith(':')) else (self.iterable))
        assert ((self.keyword=='for') and (mustbe_in=='in')), 'Bad control string %s. ' + \
                                    'Required format: "(keyword) (loopvar) (in) (iterable)"'
        
    def execute(self):
        myIterable = self.getObj(self.iterable)
        assert (hasattr(myIterable,'__iter__')), 'Loop iterable "%s" is not iterable'
        for iiiii in myIterable:
            print('Loop execution: %s=%s' % (self.loopvar,str(iiiii)))
            self.X[self.loopvar] = iiiii
            self.body.execute()
      
    
class InstructionList:
    def __init__(self,X,procname,myInput):
        self.X = X
        self.procname = procname
        self.instructions=[]
        if (isinstance(myInput,list)):  # placeholder
            self.instructions=myInput
        if ((isinstance(myInput,dict)) and ('procs' in myInput) and (procname in myInput.get('procs'))):
            # reading from procedure df. So extract df:
            myInput = myInput['procs'][procname]
        if (isinstance(myInput,pd.DataFrame)):
            self.instructions = self.parseDF(myInput)
    
    def execute(self):
        print('placeholder: now executing an instruction list:')
        assert(isinstance(self.instructions,list))
        for instruction in self.instructions:
            instruction.execute()
        print()
        
    def parseDF(self,df):
        (nest,bodies,loopEntranceTuplist)=(0,[[]],[])
        for T in df.itertuples():
            if InstructionList.isLoopEntrance(T):
                assert ((not(T.GET)) and (not(T.SET))), 'proc%s: Loop Entrance must have empty GET and SET' % self.procname
                nest+=1
                bodies.append([])
                loopEntranceTuplist.append(T)
            elif InstructionList.isLoopExit(T):
                assert ((not(T.GET)) and (not(T.SET))), 'proc%s: Loop Exit must have empty GET and SET' % self.procname
                assert (nest>0), 'proc%s: Too many Loop Exits' % self.procname
                bodies[nest-1].append(
                    Loop(X=self.X,controlString=loopEntranceTuplist.pop().TASK,body=bodies[nest],procname=self.procname))
                nest-=1
            else:
                bodies[nest].append(Statement(X=self.X,TFT=T,procname=self.procname))
        assert ((nest==0) and (not(loopEntranceTuplist))), 'proc%s: Unclosed Loop' % self.procname
        return bodies[0]
    
    @staticmethod
    def isLoopEntrance(T):
        return (T.TASK.startswith('for '))
    
    @staticmethod
    def isLoopExit(T):
        return (T.TASK.startswith('end for'))

    
class Procedure(InstructionList):
    def __init__(self,X,procname):
        assert ((procname) and isinstance(procname,str)), 'procedure name %s must be type str' % (str(procname))
        assert (('procs' in X) and (procname in X['procs'])), 'proc%s not found' % procname
        super().__init__(X,procname,X)